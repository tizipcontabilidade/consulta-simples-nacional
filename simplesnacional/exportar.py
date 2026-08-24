"""Geracao dos relatorios em Excel e CSV.

Regra combinada: CNPJ em dia entra apenas como uma linha "EM DIA" no resumo -
nada de detalhamento. A aba de ocorrencias existe so para quem tem algo a tratar.
"""
from __future__ import annotations

import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analise import ALERTA, ATENCAO, EM_DIA, ERRO, NAO_OPTANTE

CABECALHO_RESUMO = [
    "CNPJ",
    "Nome empresarial",
    "Status",
    "Situacao no Simples Nacional",
    "Situacao no SIMEI",
    "Resumo",
    "Consultado em",
]

CABECALHO_OCORRENCIAS = [
    "CNPJ",
    "Nome empresarial",
    "Status",
    "Tipo",
    "Regime",
    "Descricao",
    "Data / Periodo",
    "Comprovante",
]

_CORES = {
    ERRO: "F8CBAD",
    ALERTA: "F4B6B6",
    NAO_OPTANTE: "FDE9A9",
    ATENCAO: "FFF2CC",
    EM_DIA: "D8EFD3",
}

_FUNDO_CABECALHO = PatternFill("solid", fgColor="2F5D3A")


def _linha_resumo(dados: dict) -> list:
    return [
        dados.get("cnpj_formatado", ""),
        dados.get("nome_empresarial", ""),
        dados.get("status", ""),
        dados.get("situacao_simples", ""),
        dados.get("situacao_simei", ""),
        dados.get("resumo", ""),
        dados.get("data_consulta", ""),
    ]


def _linhas_ocorrencias(dados: dict) -> list:
    """Uma linha por ocorrencia. CNPJ em dia nao gera nenhuma."""
    if dados.get("status") == EM_DIA:
        return []

    linhas = []
    for evento in dados.get("exclusoes_futuras", []):
        linhas.append(
            [
                dados.get("cnpj_formatado", ""),
                dados.get("nome_empresarial", ""),
                dados.get("status", ""),
                "Evento futuro",
                evento.get("regime", ""),
                evento.get("descricao", ""),
                evento.get("data_efeito", ""),
                dados.get("comprovante", ""),
            ]
        )
    for ocorrencia in dados.get("historico_saidas", []):
        linhas.append(
            [
                dados.get("cnpj_formatado", ""),
                dados.get("nome_empresarial", ""),
                dados.get("status", ""),
                "Exclusao / desenquadramento anterior",
                ocorrencia.get("regime", ""),
                ocorrencia.get("detalhamento", ""),
                ocorrencia.get("periodo", ""),
                dados.get("comprovante", ""),
            ]
        )
    if not linhas:
        # Nao optante ou erro: uma unica linha explicando o motivo.
        linhas.append(
            [
                dados.get("cnpj_formatado", ""),
                dados.get("nome_empresarial", ""),
                dados.get("status", ""),
                "Situacao atual" if dados.get("status") == NAO_OPTANTE else "Falha na consulta",
                "Simples Nacional",
                dados.get("resumo", "") or dados.get("erro", ""),
                dados.get("data_consulta", ""),
                dados.get("comprovante", ""),
            ]
        )
    return linhas


def _formatar_aba(aba, cabecalho: list, larguras: list) -> None:
    for coluna, titulo in enumerate(cabecalho, start=1):
        celula = aba.cell(row=1, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _FUNDO_CABECALHO
        celula.alignment = Alignment(vertical="center")
    for coluna, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(coluna)].width = largura
    aba.freeze_panes = "A2"
    aba.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}1"


def gerar_excel(resultados: list, caminho) -> None:
    """Grava a planilha com as abas Resumo e Ocorrencias."""
    planilha = Workbook()

    resumo = planilha.active
    resumo.title = "Resumo"
    _formatar_aba(resumo, CABECALHO_RESUMO, [20, 42, 14, 40, 26, 60, 20])
    for dados in resultados:
        resumo.append(_linha_resumo(dados))
        cor = _CORES.get(dados.get("status", ""))
        if cor:
            resumo.cell(row=resumo.max_row, column=3).fill = PatternFill("solid", fgColor=cor)

    ocorrencias = planilha.create_sheet("Ocorrencias")
    _formatar_aba(ocorrencias, CABECALHO_OCORRENCIAS, [20, 38, 14, 30, 18, 60, 22, 34])
    for dados in resultados:
        for linha in _linhas_ocorrencias(dados):
            ocorrencias.append(linha)
    if ocorrencias.max_row == 1:
        ocorrencias.append(["", "", "", "Nenhuma ocorrencia: todos os CNPJs consultados estao em dia."])

    planilha.save(caminho)


def gerar_csv(resultados: list, caminho) -> None:
    with open(caminho, "w", newline="", encoding="utf-8-sig") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(CABECALHO_RESUMO)
        for dados in resultados:
            escritor.writerow(_linha_resumo(dados))


def nome_do_arquivo(prefixo: str, extensao: str) -> str:
    return f"{prefixo}-{datetime.now():%Y%m%d-%H%M%S}.{extensao}"
