"""Consulta em lote pela linha de comando - feito para rodar no Agendador de Tarefas.

Exemplos:

    python consultar.py 12345678000195 00000000000191
    python consultar.py --arquivo clientes.txt --formato xlsx
    python consultar.py --arquivo clientes.xlsx --somente-mudancas --silencioso

Codigos de saida (uteis para o Agendador e para scripts):
    0  nada a tratar
    1  ha CNPJ com ocorrencia (ou mudanca, com --somente-mudancas)
    2  erro de uso / nenhum CNPJ valido
    3  falha na execucao (navegador, portal fora do ar, etc.)
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from simplesnacional import config, exportar, historico, lote
from simplesnacional.analise import EM_DIA


def _ler_arquivo(caminho: Path) -> lote.Leitura:
    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        planilha = load_workbook(caminho, read_only=True, data_only=True)
        textos = []
        for aba in planilha.worksheets:
            for linha in aba.iter_rows(values_only=True):
                textos += [str(c) for c in linha if c is not None]
        return lote.ler("\n".join(textos))

    for codificacao in ("utf-8-sig", "latin-1"):
        try:
            return lote.ler(caminho.read_text(encoding=codificacao))
        except UnicodeDecodeError:
            continue
    return lote.Leitura()


def _argumentos():
    analisador = argparse.ArgumentParser(
        prog="consultar",
        description="Consulta CNPJs no portal do Simples Nacional e relata ocorrencias.",
    )
    analisador.add_argument("cnpjs", nargs="*", help="CNPJs a consultar (com ou sem pontuacao)")
    analisador.add_argument("-a", "--arquivo", type=Path, help="arquivo .txt, .csv ou .xlsx com os CNPJs")
    analisador.add_argument(
        "-f", "--formato", default="xlsx", choices=["xlsx", "csv", "json", "nenhum"],
        help="formato do relatorio gravado (padrao: xlsx)",
    )
    analisador.add_argument("-s", "--saida", type=Path, help="pasta onde gravar o relatorio")
    analisador.add_argument(
        "--somente-mudancas", action="store_true",
        help="relata apenas os CNPJs cujo retrato mudou desde a consulta anterior",
    )
    analisador.add_argument(
        "--sem-historico", action="store_true",
        help="nao le nem grava o historico de comparacao",
    )
    analisador.add_argument(
        "--oculto", action="store_true",
        help="roda com o navegador escondido (se o portal pedir captcha, o CNPJ falha)",
    )
    analisador.add_argument("--silencioso", action="store_true", help="imprime so o resumo final")
    return analisador.parse_args()


def main() -> int:
    argumentos = _argumentos()

    leitura = lote.ler(" ".join(argumentos.cnpjs))
    if argumentos.arquivo:
        if not argumentos.arquivo.exists():
            print(f"arquivo nao encontrado: {argumentos.arquivo}", file=sys.stderr)
            return 2
        leitura.mesclar(_ler_arquivo(argumentos.arquivo))

    # O que foi lido e nao virou consulta sai antes do lote comecar: numero que
    # entrou tem de bater com numero que saiu.
    if leitura.descartados:
        print(
            f"{leitura.total_lido} entrada(s) lida(s); "
            f"{len(leitura.descartados)} nao sera(ao) consultada(s):",
            file=sys.stderr,
        )
        for descartado in leitura.descartados:
            print(f"  {descartado.bruto}  ->  {descartado.motivo}", file=sys.stderr)
        print(file=sys.stderr)

    cnpjs = leitura.cnpjs
    if not cnpjs:
        print("nenhum CNPJ valido informado", file=sys.stderr)
        return 2

    def progresso(execucao):
        if not argumentos.silencioso:
            print(f"[{execucao.percentual:3d}%] {execucao.mensagem}", flush=True)

    execucao = lote.executar(
        cnpjs,
        execucao=lote.Execucao(cnpjs=cnpjs, descartados=leitura.descartados),
        visivel=not argumentos.oculto,
        usar_historico=not argumentos.sem_historico,
        ao_progredir=progresso,
    )

    if execucao.erro_fatal:
        print(f"falha: {execucao.erro_fatal}", file=sys.stderr)
        return 3

    itens = [i.como_dicionario() for i in execucao.resultados_ordenados()]
    if argumentos.somente_mudancas:
        itens = [i for i in itens if i.get("situacao_historico") == historico.MUDOU]

    relevantes = [i for i in itens if i.get("status") != EM_DIA]
    descartados = [d.como_dicionario() for d in execucao.descartados]

    print()
    print(f"{execucao.total} CNPJ(s) consultado(s) em {datetime.now():%d/%m/%Y %H:%M}")
    for status, quantidade in sorted(execucao.contagem().items()):
        print(f"  {status:<12} {quantidade}")
    if not argumentos.sem_historico:
        print(f"  {'MUDOU':<12} {len(execucao.com_mudanca())}")

    if relevantes:
        print()
        titulo = "Mudancas desde a ultima consulta:" if argumentos.somente_mudancas else "Ocorrencias:"
        print(titulo)
        for dados in relevantes:
            print(f"  [{dados.get('status')}] {dados.get('cnpj_formatado')} "
                  f"{dados.get('nome_empresarial', '')}")
            print(f"      {dados.get('resumo', '')}")
            for mudanca in dados.get("mudancas", []):
                if mudanca["tipo"] != historico.PRIMEIRA:
                    print(f"      * {mudanca['tipo']}: {mudanca['descricao']}")

    if argumentos.formato != "nenhum" and itens:
        pasta = argumentos.saida or config.PASTA_SAIDA
        pasta.mkdir(parents=True, exist_ok=True)
        carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
        caminho = pasta / f"simples-nacional-{carimbo}.{argumentos.formato}"
        if argumentos.formato == "xlsx":
            exportar.gerar_excel(itens, caminho, descartados)
        elif argumentos.formato == "csv":
            exportar.gerar_csv(itens, caminho, descartados)
        else:
            lote.salvar_json(execucao, caminho)
        print()
        print(f"relatorio: {caminho}")

    return 1 if relevantes else 0


if __name__ == "__main__":
    sys.exit(main())
