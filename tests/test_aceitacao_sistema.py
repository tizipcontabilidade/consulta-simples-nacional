"""System Acceptance Test: o sistema inteiro, de ponta a ponta.

Entra pela interface web como um usuario entraria, passa pelo lote, pela
classificacao, pelo historico e sai nos tres formatos de relatorio. So o
navegador e dublado.
"""
from __future__ import annotations

import io
import json

import pytest
from openpyxl import load_workbook

import app as aplicacao
from simplesnacional.analise import ALERTA, EM_DIA, NAO_OPTANTE

from . import fixturas

TRES_CNPJS = f"{fixturas.CNPJ_COM_EVENTOS}\n{fixturas.CNPJ_NAO_OPTANTE}\n{fixturas.CNPJ_EM_DIA}"


@pytest.fixture
def cliente(pastas_temporarias, sessao_falsa, monkeypatch):
    monkeypatch.setattr(aplicacao, "_atual", None)
    monkeypatch.setattr(aplicacao, "_thread", None)
    aplicacao.app.config["TESTING"] = True
    with aplicacao.app.test_client() as teste:
        yield teste


def _rodar(cliente, corpo=None, arquivo=None):
    dados = {"cnpjs": corpo or "", "visivel": "on"}
    if arquivo:
        dados["arquivo"] = arquivo
    resposta = cliente.post("/consultar", data=dados, content_type="multipart/form-data")
    if aplicacao._thread:
        aplicacao._thread.join(timeout=30)
    return resposta


def test_fluxo_completo_pela_interface(cliente):
    resposta = _rodar(cliente, TRES_CNPJS)
    assert resposta.status_code == 302
    assert "/andamento" in resposta.headers["Location"]

    estado = cliente.get("/api/estado").get_json()
    assert estado["concluido"] is True
    assert estado["total"] == 3
    assert estado["processados"] == 3
    assert estado["contagem"] == {ALERTA: 1, NAO_OPTANTE: 1, EM_DIA: 1}

    # A tela abre no grupo mais grave que tenha algo: quem chega no resultado
    # quer ver o que precisa de acao, nao a lista dos que estao em dia.
    pagina = cliente.get("/resultado").get_data(as_text=True)
    assert "11.222.333/0001-81" in pagina
    assert "MARCENARIA MODELO LTDA" in pagina
    assert "Exclusão de Ofício - Débitos" in pagina
    assert "01/01/2027" in pagina

    # Os demais grupos ficam a um clique, com a contagem no proprio botao.
    for grupo in (ALERTA, NAO_OPTANTE, EM_DIA):
        assert f"grupo={grupo.replace(' ', '+')}" in pagina

    em_dia = cliente.get(f"/resultado?grupo={EM_DIA}").get_data(as_text=True)
    assert "PADARIA EXEMPLO ME" in em_dia


def test_ordem_dos_grupos_e_por_gravidade(cliente):
    """Os botoes de filtro seguem a gravidade, e a tela abre no primeiro."""
    _rodar(cliente, TRES_CNPJS)
    pagina = cliente.get("/resultado").get_data(as_text=True)

    posicoes = [pagina.index(f"grupo={g.replace(' ', '+')}") for g in (ALERTA, NAO_OPTANTE, EM_DIA)]
    assert posicoes == sorted(posicoes)
    assert "11.222.333/0001-81" in pagina, "abre no grupo mais grave"


def test_um_grupo_por_vez_para_lote_grande(cliente):
    """O motivo do filtro existir: mil CNPJs numa pagina so era inutilizavel."""
    _rodar(cliente, TRES_CNPJS)

    alerta = cliente.get(f"/resultado?grupo={ALERTA}").get_data(as_text=True)

    assert "11.222.333/0001-81" in alerta
    assert "PADARIA EXEMPLO ME" not in alerta, "o grupo em dia nao entra aqui"
    assert "00.000.000/0001-91" not in alerta


def test_planilha_tem_resumo_completo_e_ocorrencias_filtradas(cliente):
    _rodar(cliente, TRES_CNPJS)
    resposta = cliente.get("/baixar/xlsx")
    assert resposta.status_code == 200

    planilha = load_workbook(io.BytesIO(resposta.data))
    assert planilha.sheetnames == ["Resumo", "Ocorrencias", "Nao consultados"]

    resumo = list(planilha["Resumo"].iter_rows(values_only=True))
    assert len(resumo) == 4                     # cabecalho + 3 CNPJs
    assert [linha[2] for linha in resumo[1:]] == [ALERTA, NAO_OPTANTE, EM_DIA]

    ocorrencias = list(planilha["Ocorrencias"].iter_rows(values_only=True))
    cnpjs_com_ocorrencia = {linha[0] for linha in ocorrencias[1:]}
    assert "11.444.777/0001-61" not in cnpjs_com_ocorrencia   # o que esta em dia
    assert "11.222.333/0001-81" in cnpjs_com_ocorrencia


def test_csv_traz_todos_os_cnpjs(cliente):
    _rodar(cliente, TRES_CNPJS)
    texto = cliente.get("/baixar/csv").get_data(as_text=True)

    assert texto.count("\n") >= 4
    for formatado in ("11.222.333/0001-81", "00.000.000/0001-91", "11.444.777/0001-61"):
        assert formatado in texto


def test_json_traz_a_estrutura_completa(cliente):
    _rodar(cliente, TRES_CNPJS)
    dados = json.loads(cliente.get("/baixar/json").get_data(as_text=True))

    assert dados["total"] == 3
    alerta = next(r for r in dados["resultados"] if r["status"] == ALERTA)
    assert alerta["eventos_futuros_sn"][0]["data_efeito"] == "01/01/2027"
    assert alerta["nome_empresarial"] == "MARCENARIA MODELO LTDA"
    assert alerta["comprovante"].endswith(".html")


def test_comprovante_fica_disponivel_para_abrir(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    dados = json.loads(cliente.get("/baixar/json").get_data(as_text=True))
    nome = dados["resultados"][0]["comprovante"]

    resposta = cliente.get(f"/comprovante/{nome}")
    assert resposta.status_code == 200
    assert "Exclus" in resposta.get_data(as_text=True)


def test_consulta_por_arquivo_enviado(cliente):
    conteudo = f"# lista da carteira\n{fixturas.CNPJ_COM_EVENTOS}\n{fixturas.CNPJ_EM_DIA}\n".encode()
    _rodar(cliente, arquivo=(io.BytesIO(conteudo), "clientes.txt"))

    estado = cliente.get("/api/estado").get_json()
    assert estado["total"] == 2


def test_lista_vazia_avisa_em_vez_de_rodar(cliente):
    resposta = cliente.post("/consultar", data={"cnpjs": "sem nada aqui"})

    assert resposta.status_code == 200
    assert "Nenhum CNPJ valido encontrado" in resposta.get_data(as_text=True)


def test_cnpjs_repetidos_sao_consultados_uma_vez_so(cliente):
    repetido = f"{fixturas.CNPJ_EM_DIA}\n11.444.777/0001-61\n{fixturas.CNPJ_EM_DIA}"
    _rodar(cliente, repetido)

    assert cliente.get("/api/estado").get_json()["total"] == 1
