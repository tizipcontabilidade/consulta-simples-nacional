"""User Acceptance Test: cada criterio pedido pelo usuario, na ordem em que foi pedido.

Os nomes dos testes repetem o pedido original de proposito - e por eles que se
confere se o sistema entrega o que foi combinado.
"""
from __future__ import annotations

import io
import json

import pytest
from openpyxl import load_workbook

import app as aplicacao
from simplesnacional import historico, lote
from simplesnacional.analise import ALERTA, EM_DIA

from . import fixturas


@pytest.fixture
def cliente(pastas_temporarias, sessao_falsa, monkeypatch):
    monkeypatch.setattr(aplicacao, "_atual", None)
    monkeypatch.setattr(aplicacao, "_thread", None)
    with aplicacao.app.test_client() as teste:
        yield teste


def _rodar(cliente, corpo):
    cliente.post("/consultar", data={"cnpjs": corpo, "visivel": "on"})
    if aplicacao._thread:
        aplicacao._thread.join(timeout=30)


# "pesquisa de CNPJs unicos ou em lote"
def test_consulta_um_cnpj_sozinho(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    assert cliente.get("/api/estado").get_json()["total"] == 1


def test_consulta_em_lote(cliente):
    lista = "\n".join([fixturas.CNPJ_COM_EVENTOS, fixturas.CNPJ_EM_DIA, fixturas.CNPJ_NAO_OPTANTE])
    _rodar(cliente, lista)
    assert cliente.get("/api/estado").get_json()["total"] == 3


# "informe se existe desenquadramento do simples nacional"
def test_informa_desenquadramento_e_exclusao_anterior(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    pagina = cliente.get("/resultado").get_data(as_text=True)

    assert "Exclusoes / desenquadramentos anteriores" in pagina
    assert "Excluída por Ato Administrativo praticado pela Receita Federal do Brasil" in pagina
    assert "Desenquadrada por Comunicação Obrigatória do Contribuinte" in pagina
    assert "01/01/2024 a 31/12/2024" in pagina


# "aviso de eventos futuros (Como Exclusao de Oficio)"
def test_avisa_evento_futuro_de_exclusao_de_oficio(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    pagina = cliente.get("/resultado").get_data(as_text=True)

    assert "Eventos futuros" in pagina
    assert "Exclusão de Ofício - Débitos" in pagina
    assert "01/01/2027" in pagina
    assert "ALERTA" in pagina


# "e qualquer outra informacao referente ao cnpj consultado"
def test_traz_as_demais_informacoes_do_cnpj(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    dados = json.loads(cliente.get("/baixar/json").get_data(as_text=True))["resultados"][0]

    for campo in (
        "cnpj_formatado", "nome_empresarial", "data_consulta",
        "situacao_simples", "optante_desde", "situacao_simei",
        "periodos_anteriores_sn", "periodos_anteriores_simei",
        "eventos_futuros_sn", "eventos_futuros_simei", "mei_transportador",
    ):
        assert campo in dados, f"campo ausente no relatorio: {campo}"
    assert dados["nome_empresarial"] == "MARCENARIA MODELO LTDA"
    assert dados["situacao_simples"].startswith("Optante")


# "se o cnpj estiver em dia, nao gerar linha detalhada"
def test_cnpj_em_dia_aparece_so_no_resumo(cliente):
    _rodar(cliente, f"{fixturas.CNPJ_EM_DIA}\n{fixturas.CNPJ_COM_EVENTOS}")

    # O grupo de ocorrencias nao traz quem esta em dia; o grupo EM DIA traz,
    # e so como linha de resumo.
    ocorrencias = cliente.get("/resultado?grupo=ALERTA").get_data(as_text=True)
    assert "11.444.777/0001-61" not in ocorrencias

    em_dia = cliente.get("/resultado?grupo=EM+DIA").get_data(as_text=True)
    assert "11.444.777/0001-61" in em_dia
    assert "Eventos futuros" not in em_dia, "em dia nao ganha detalhamento"

    planilha = load_workbook(io.BytesIO(cliente.get("/baixar/xlsx").data))
    resumo = [linha[0] for linha in planilha["Resumo"].iter_rows(min_row=2, values_only=True)]
    ocorrencias = [linha[0] for linha in planilha["Ocorrencias"].iter_rows(min_row=2, values_only=True)]
    assert "11.444.777/0001-61" in resumo
    assert "11.444.777/0001-61" not in ocorrencias


def test_cnpj_em_dia_nao_consome_espaco_com_comprovante(cliente):
    _rodar(cliente, fixturas.CNPJ_EM_DIA)
    dados = json.loads(cliente.get("/baixar/json").get_data(as_text=True))["resultados"][0]

    assert dados["status"] == EM_DIA
    assert dados["comprovante"] == ""


# "comparacao com a consulta anterior para avisar so o que mudou"
def test_avisa_o_que_mudou_desde_a_ultima_consulta(pastas_temporarias, sessao_falsa):
    # Primeira rodada: nada com que comparar.
    primeira = lote.executar([fixturas.CNPJ_COM_EVENTOS], visivel=False)
    assert primeira.itens[0].situacao_historico == historico.PRIMEIRA
    assert primeira.com_mudanca() == []

    # Segunda rodada, sem novidade no portal: nada muda.
    segunda = lote.executar([fixturas.CNPJ_COM_EVENTOS], visivel=False)
    assert segunda.itens[0].situacao_historico == historico.SEM_MUDANCA

    # O portal passa a acusar um evento novo.
    estado = historico.carregar()
    estado[fixturas.CNPJ_COM_EVENTOS]["eventos"] = []
    historico.gravar(estado)

    terceira = lote.executar([fixturas.CNPJ_COM_EVENTOS], visivel=False)
    item = terceira.itens[0]
    assert item.situacao_historico == historico.MUDOU
    assert len(terceira.com_mudanca()) == 1
    assert item.mudancas[0].tipo == "Novo evento futuro"
    assert "01/01/2027" in item.mudancas[0].descricao


def test_tela_de_resultado_destaca_o_que_mudou(cliente):
    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    estado = historico.carregar()
    estado[fixturas.CNPJ_COM_EVENTOS]["situacao_simples"] = "NÃO optante pelo Simples Nacional"
    historico.gravar(estado)

    _rodar(cliente, fixturas.CNPJ_COM_EVENTOS)
    # O que mudou vira o primeiro filtro da tela, e e nele que ela abre.
    pagina = cliente.get("/resultado").get_data(as_text=True)

    assert "grupo=MUDOU" in pagina
    assert "Situacao no Simples Nacional" in pagina


# "linha de comando para agendar" - criterio da segunda rodada de pedidos
def test_relatorio_so_com_mudancas_fica_vazio_quando_nada_mudou(pastas_temporarias, sessao_falsa):
    lote.executar([fixturas.CNPJ_COM_EVENTOS], visivel=False)
    segunda = lote.executar([fixturas.CNPJ_COM_EVENTOS], visivel=False)

    apenas_mudancas = [
        i for i in segunda.itens if i.situacao_historico == historico.MUDOU
    ]
    assert apenas_mudancas == []
    assert segunda.itens[0].veredito.status == ALERTA  # continua sendo ocorrencia
