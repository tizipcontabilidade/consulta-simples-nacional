"""Interface web local para a consulta em lote do Simples Nacional.

Suba com:  python app.py     e abra http://127.0.0.1:5000
"""
from __future__ import annotations

import io
import os
import socket
import threading
import time
import webbrowser
from datetime import datetime

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)

from simplesnacional import atualizacao, config, exportar, historico, lote
from simplesnacional.analise import EM_DIA, ORDEM, ROTULOS
from simplesnacional.versao import VERSAO

app = Flask(
    __name__,
    template_folder=str(config.RAIZ_CODIGO / "templates"),
    static_folder=str(config.RAIZ_CODIGO / "static"),
)

# Um lote por vez: a consulta depende de uma janela de navegador unica.
_atual: lote.Execucao | None = None
_thread: threading.Thread | None = None
_trava = threading.Lock()

# O sistema roda sem janela de console: quem manda encerrar e a propria
# interface. A pagina envia um sinal de vida a cada poucos segundos; se a aba
# for fechada (ou o navegador todo), o sinal para e o servidor se encerra
# sozinho - mas nunca no meio de um lote.
_ultimo_sinal = time.monotonic()
_encerrando = False
INTERVALO_SINAL = 5                            # ritmo da verificacao
TOLERANCIA_SINAL = config.TOLERANCIA_SINAL     # silencio maior = ninguem olhando


# Manifesto de versao relido de tempos em tempos. E leitura de arquivo local
# (pasta do Drive sincronizada), entao custa pouco - mas nao a cada clique.
_atualizacao = atualizacao.Atualizacao()
_atualizacao_vista_em = 0.0


def _ha_lote_ativo() -> bool:
    return _thread is not None and _thread.is_alive()


def _checar_atualizacao() -> atualizacao.Atualizacao:
    global _atualizacao, _atualizacao_vista_em
    agora = time.monotonic()
    if agora - _atualizacao_vista_em >= config.INTERVALO_ATUALIZACAO:
        _atualizacao_vista_em = agora
        _atualizacao = atualizacao.verificar()
    return _atualizacao


@app.context_processor
def _contexto_padrao():
    """Versao e aviso de atualizacao ficam disponiveis em todas as telas."""
    return {"versao": VERSAO, "atualizacao": _checar_atualizacao()}


def _encerrar_processo(atraso: float = 0.6) -> None:
    """Derruba o servidor local. os._exit e o caminho confiavel aqui:
    o servidor de desenvolvimento do Flask nao tem parada limpa por fora
    de uma requisicao, e nao ha estado em memoria para preservar."""
    global _encerrando
    _encerrando = True
    threading.Timer(atraso, lambda: os._exit(0)).start()


def _vigiar_interface() -> None:
    """Encerra o sistema quando ninguem mais tem a interface aberta."""
    while not _encerrando:
        time.sleep(INTERVALO_SINAL)
        if _ha_lote_ativo():
            continue  # lote rodando: segura o encerramento ate terminar
        if time.monotonic() - _ultimo_sinal > TOLERANCIA_SINAL:
            _encerrar_processo(0.1)
            return


def _porta_ocupada(porta: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as teste:
        teste.settimeout(0.4)
        return teste.connect_ex(("127.0.0.1", porta)) == 0


# --------------------------------------------------------------------- leitura
def _cnpjs_de_planilha(arquivo) -> lote.Leitura:
    """Le CNPJs de um .xlsx enviado (qualquer coluna, qualquer aba)."""
    from openpyxl import load_workbook

    planilha = load_workbook(io.BytesIO(arquivo.read()), read_only=True, data_only=True)
    textos = []
    for aba in planilha.worksheets:
        for linha in aba.iter_rows(values_only=True):
            for celula in linha:
                if celula is not None:
                    textos.append(str(celula))
    return lote.ler("\n".join(textos))


def _cnpjs_do_pedido() -> lote.Leitura:
    """Junta o que foi colado com o que veio de arquivo, guardando o descarte."""
    leitura = lote.ler(request.form.get("cnpjs", ""))

    arquivo = request.files.get("arquivo")
    if arquivo and arquivo.filename:
        nome = arquivo.filename.lower()
        if nome.endswith((".xlsx", ".xlsm")):
            leitura.mesclar(_cnpjs_de_planilha(arquivo))
        else:
            bruto = arquivo.read()
            for codificacao in ("utf-8-sig", "latin-1"):
                try:
                    leitura.mesclar(lote.ler(bruto.decode(codificacao)))
                    break
                except UnicodeDecodeError:
                    continue

    return leitura


# ---------------------------------------------------------------------- rotas
@app.get("/")
def inicio():
    return render_template("index.html", execucao=_atual)


@app.post("/consultar")
def consultar():
    global _atual, _thread

    leitura = _cnpjs_do_pedido()
    if not leitura.cnpjs:
        return render_template(
            "index.html",
            execucao=None,
            aviso="Nenhum CNPJ valido encontrado. Cole a lista ou envie um arquivo .txt, .csv ou .xlsx.",
            descartados=leitura.descartados,
        )

    with _trava:
        if _thread and _thread.is_alive():
            return redirect(url_for("andamento"))

        _atual = lote.Execucao(cnpjs=leitura.cnpjs, descartados=leitura.descartados)
        visivel = request.form.get("visivel") == "on"
        todos_comprovantes = request.form.get("comprovante_todos") == "on"

        def rodar(execucao=_atual):
            lote.executar(
                execucao.cnpjs,
                execucao=execucao,
                visivel=visivel,
                salvar_comprovante_em_dia=todos_comprovantes,
            )

        _thread = threading.Thread(target=rodar, daemon=True)
        _thread.start()

    return redirect(url_for("andamento"))


@app.get("/andamento")
def andamento():
    if _atual is None:
        return redirect(url_for("inicio"))
    return render_template("andamento.html", execucao=_atual)


@app.get("/api/estado")
def api_estado():
    if _atual is None:
        return jsonify({"vazio": True})
    return jsonify(_atual.como_dicionario())


@app.post("/cancelar")
def cancelar():
    if _atual is not None:
        _atual.cancelado = True
        _atual.mensagem = "Cancelando apos a consulta em andamento..."
    return redirect(url_for("andamento"))


@app.get("/resultado")
def resultado():
    if _atual is None:
        return redirect(url_for("inicio"))

    itens = [i.como_dicionario() for i in _atual.resultados_ordenados()]
    com_ocorrencia = [i for i in itens if i.get("status") != EM_DIA]
    em_dia = [i for i in itens if i.get("status") == EM_DIA]
    mudaram = [i for i in itens if i.get("situacao_historico") == historico.MUDOU]
    return render_template(
        "resultado.html",
        execucao=_atual,
        com_ocorrencia=com_ocorrencia,
        em_dia=em_dia,
        mudaram=mudaram,
        rotulos=ROTULOS,
        ordem=ORDEM,
        descartados=_atual.descartados,
    )


@app.get("/baixar/<formato>")
def baixar(formato: str):
    if _atual is None:
        return redirect(url_for("inicio"))

    itens = [i.como_dicionario() for i in _atual.resultados_ordenados()]
    descartados = [d.como_dicionario() for d in _atual.descartados]
    config.PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")

    if formato == "xlsx":
        caminho = config.PASTA_SAIDA / f"simples-nacional-{carimbo}.xlsx"
        exportar.gerar_excel(itens, caminho, descartados)
    elif formato == "csv":
        caminho = config.PASTA_SAIDA / f"simples-nacional-{carimbo}.csv"
        exportar.gerar_csv(itens, caminho, descartados)
    elif formato == "json":
        caminho = config.PASTA_SAIDA / f"simples-nacional-{carimbo}.json"
        lote.salvar_json(_atual, caminho)
    else:
        abort(404)

    return send_file(caminho, as_attachment=True)


@app.post("/sinal")
def sinal():
    """Batida de coracao da interface: diz que ainda tem gente com a tela aberta."""
    global _ultimo_sinal
    _ultimo_sinal = time.monotonic()
    return ("", 204)


@app.get("/api/atualizacao")
def api_atualizacao():
    return jsonify(_checar_atualizacao().como_dicionario())


@app.post("/atualizar")
def atualizar():
    """Confere o instalador publicado e o abre, depois se encerra.

    Um lote em andamento sempre vence: atualizar no meio de uma consulta
    perderia o trabalho, que e exatamente o que este sistema existe para evitar.
    """
    if _ha_lote_ativo():
        return render_template(
            "andamento.html",
            execucao=_atual,
            aviso="Ha um lote em andamento. A atualizacao espera ele terminar.",
        )

    disponivel = _checar_atualizacao()
    if not disponivel.instalavel:
        return render_template(
            "index.html",
            execucao=None,
            aviso=disponivel.problema or "Nao ha atualizacao disponivel agora.",
        )

    problema = atualizacao.abrir_instalador(disponivel)
    if problema:
        return render_template("index.html", execucao=None, aviso=problema)

    # O instalador precisa substituir o executavel que esta rodando agora.
    _encerrar_processo(2.0)
    return render_template("atualizando.html", atualizacao=disponivel)


@app.post("/encerrar")
def encerrar():
    if _ha_lote_ativo():
        return redirect(url_for("andamento"))
    _encerrar_processo()
    return render_template("encerrado.html")


@app.get("/comprovante/<nome>")
def comprovante(nome: str):
    return send_from_directory(config.PASTA_COMPROVANTES, nome)


def iniciar_servidor(porta: int = 5000, abrir_navegador: bool = True) -> None:
    """Sobe a interface local e abre o navegador padrao na tela inicial.

    Se o sistema ja estiver rodando (atalho clicado duas vezes), apenas traz a
    tela de volta em vez de subir um segundo servidor e falhar na porta.
    """
    config.preparar_pastas()
    endereco = f"http://127.0.0.1:{porta}"

    if _porta_ocupada(porta):
        if abrir_navegador:
            webbrowser.open(endereco)
        print("O sistema ja estava aberto; trouxe a tela de volta.")
        return

    if abrir_navegador:
        threading.Timer(1.2, lambda: webbrowser.open(endereco)).start()
    threading.Thread(target=_vigiar_interface, daemon=True).start()

    print(f"Consulta Simples Nacional rodando em {endereco}")
    app.run(host="127.0.0.1", port=porta, debug=False)


if __name__ == "__main__":
    iniciar_servidor()
